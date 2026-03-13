import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
import random
import io
import zipfile

# 页面配置
st.set_page_config(page_title="RandomSCC 自动化工具", layout="wide")
st.title("📦 RandomSCC 数据填充工具")
st.markdown("""
### 使用说明：
1. 依次上传三个必要的 Excel 文件。
2. 系统检查文件无误后，点击“生成”按钮。
3. 处理完成后，点击下载生成的 ZIP 压缩包。
""")

# --- 第一部分：文件上传界面 ---
col1, col2, col3 = st.columns(3)

with col1:
    st.subheader("1. 整柜数据")
    ci_file = st.file_uploader("上传 containerinformation.xlsx", type=["xlsx"], key="ci")

with col2:
    st.subheader("2. ICS 表头")
    tmpl_file = st.file_uploader("上传 icstemplate.xlsx", type=["xlsx"], key="tmpl")

with col3:
    st.subheader("3. RSC 源数据")
    sc_file = st.file_uploader("上传 realsc.xlsx", type=["xlsx"], key="sc")

# --- 第二部分：核心处理逻辑 ---
if ci_file and tmpl_file and sc_file:
    st.success("✅ 三个文件已全部上传，可以开始处理！")
    
    if st.button("🚀 批量生成并打包下载"):
        try:
            # 1. 处理 realsc 数据池
            df_realsc = pd.read_excel(sc_file, header=None)
            df_realsc.dropna(how='all', inplace=True)
            realsc_data = df_realsc.values.tolist()
            # 每4行一组
            realsc_groups = [realsc_data[i:i+4] for i in range(0, len(realsc_data), 4) if len(realsc_data[i:i+4]) == 4]

            # 2. 处理 containerinformation
            df_ci = pd.read_excel(ci_file)
            if '单号' not in df_ci.columns:
                st.error("错误：containerinformation 文件中未找到 '单号' 列！")
                st.stop()
            
            df_ci['单号'] = df_ci['单号'].ffill()
            grouped = df_ci.groupby('单号')

            # 3. 准备 ZIP 压缩包容器
            zip_buffer = io.BytesIO()
            
            # 读取模板的基础字节，方便重复加载
            template_bytes = tmpl_file.read()

            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                # 样式定义
                custom_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                custom_font = Font(name='微软雅黑', size=11)

                progress_bar = st.progress(0)
                status_text = st.empty()
                total_groups = len(grouped)

                for idx, (order_no, group) in enumerate(grouped):
                    status_text.text(f"正在处理单号: {order_no} ({idx+1}/{total_groups})")
                    
                    # 每次从内存字节中重新加载模板
                    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
                    ws = wb.active 
                    
                    # --- 填充逻辑 ---
                    ws['B5'] = str(order_no).upper()
                    f130_val = ws['F130'].value

                    if len(group) == 1:
                        row_data = group.iloc[0]
                        ws['B8'], ws['B9'] = row_data['件数'], row_data['重量(KGS)']
                        ws['A130'], ws['B130'] = str(row_data['HS CODE']).upper(), str(row_data['品名']).upper()
                        ws['C130'], ws['E130'] = row_data['件数'], row_data['重量(KGS)']
                    else:
                        ws['B8'], ws['B9'] = group['件数'].sum(), group['重量(KGS)'].sum()
                        for i, (_, row_data) in enumerate(group.iterrows()):
                            curr_row = 130 + i
                            ws.cell(row=curr_row, column=1, value=str(row_data['HS CODE']).upper())
                            ws.cell(row=curr_row, column=2, value=str(row_data['品名']).upper())
                            ws.cell(row=curr_row, column=3, value=row_data['件数'])
                            ws.cell(row=curr_row, column=4, value="PK-PACKAGE")
                            ws.cell(row=curr_row, column=5, value=row_data['重量(KGS)'])
                            ws.cell(row=curr_row, column=6, value=f130_val)

                    # --- 随机填充 realsc ---
                    if realsc_groups:
                        chosen_sc = random.choice(realsc_groups)
                        target_rows = [14, 15, 18, 19]
                        for r_idx, data_row in enumerate(chosen_sc):
                            for c_idx, value in enumerate(data_row, start=3):
                                if c_idx <= 8:
                                    final_val = str(value).upper() if isinstance(value, str) else value
                                    ws.cell(row=target_rows[r_idx], column=c_idx, value=final_val)

                    # --- 全表格式化 ---
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = custom_alignment
                            cell.font = custom_font
                            if isinstance(cell.value, str):
                                cell.value = cell.value.upper()

                    # 保存当前生成的 Excel 到 ZIP
                    file_stream = io.BytesIO()
                    wb.save(file_stream)
                    zip_file.writestr(f"{order_no}.xlsx", file_stream.getvalue())
                    
                    progress_bar.progress((idx + 1) / total_groups)

            status_text.text("✅ 所有文件处理完成！")
            
            # 4. 提供 ZIP 下载按钮
            st.download_button(
                label="📥 点击下载所有生成的文件 (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="processed_results.zip",
                mime="application/x-zip-compressed"
            )
        except Exception as e:
            st.error(f"处理过程中发生错误: {e}")
else:
    st.info("💡 请在上方上传所有三个 Excel 文件以解锁生成按钮。")
